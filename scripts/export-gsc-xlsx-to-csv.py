# -*- coding: utf-8 -*-
"""Convert GSC Thai-UI xlsx exports in docs/gsc/ to CSV for analyze-gsc-query-page.mjs"""
import csv
import json
from datetime import date, datetime
from pathlib import Path
import openpyxl

OUT = Path("docs/gsc")


def cell(v):
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return v


def write_sheet(wb, sheet_name, out_name):
    if sheet_name not in wb.sheetnames:
        return False
    ws = wb[sheet_name]
    rows = [[cell(x) for x in r] for r in ws.iter_rows(values_only=True)]
    with (OUT / out_name).open("w", encoding="utf-8-sig", newline="") as f:
        csv.writer(f).writerows(rows)
    return True


def main():
    perf = next(OUT.glob("*Performance*.xlsx"), None)
    if not perf:
        raise SystemExit("No Performance xlsx in docs/gsc/")
    wb = openpyxl.load_workbook(perf, read_only=True, data_only=True)
    write_sheet(wb, "ข้อความค้นหา", "queries-3m.csv")
    write_sheet(wb, "หน้า", "pages-3m.csv")
    write_sheet(wb, "แผนผัง", "daily-3m.csv")
    if "ตัวกรอง" in wb.sheetnames:
        filters = {
            str(r[0]).strip(): str(r[1]).strip()
            for r in wb["ตัวกรอง"].iter_rows(values_only=True)
            if r and r[0]
        }
        (OUT / "filters-3m.json").write_text(json.dumps(filters, ensure_ascii=False, indent=2), encoding="utf-8")
    wb.close()

    cov = next(OUT.glob("*Coverage-Valid*.xlsx"), None)
    if cov:
        wb2 = openpyxl.load_workbook(cov, read_only=True, data_only=True)
        write_sheet(wb2, "ตาราง", "coverage-valid-urls.csv")
        wb2.close()

    cov2 = next((p for p in OUT.glob("*Coverage*.xlsx") if "Valid" not in p.name), None)
    if cov2:
        wb3 = openpyxl.load_workbook(cov2, read_only=True, data_only=True)
        write_sheet(wb3, "ปัญหาร้ายแรง", "coverage-issues.csv")
        wb3.close()

    print("OK: exported CSVs beside xlsx in docs/gsc/")


if __name__ == "__main__":
    main()
